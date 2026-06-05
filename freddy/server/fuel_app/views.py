from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST, require_http_methods
from rest_framework import status as drf_status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from authentication.models import ROLE_NGO_ADMIN, ROLE_COMPANY_MANAGER, ROLE_STATION_AGENT, User
from fuel_app.forms import (
    AgentForm, ChurchForm, DisbursementForm, FuelStationForm, FuelTypeForm,
    ParentCompanyForm, StationTargetForm, TransactionFilterForm, TransactionStatusForm,
)
from fuel_app.models import (
    Church, Disbursement, Driver, ExchangeRateCache, FuelStation, FuelType,
    ParentCompany, StationTarget, Transaction, TransactionAuditLog,
)
from fuel_app.permissions import IsNGOAdmin, IsStationAgent
from fuel_app.serializers import (
    ChurchSerializer, FuelStationSerializer, FuelTypeSerializer,
    TransactionAuditLogSerializer, TransactionCreateSerializer,
    TransactionSerializer, TransactionStatusUpdateSerializer,
)
from fuel_app.services import get_usd_to_cdf_rate, record_audit_log


# ─── helpers ──────────────────────────────────────────────────────────────────

def _htmx(request):
    return request.headers.get("HX-Request") == "true"


def _tx_queryset(request):
    qs = Transaction.objects.select_related(
        "station__company", "church", "agent", "fuel_type"
    )
    f = TransactionFilterForm(request.GET)
    if f.is_valid():
        d = f.cleaned_data
        if d.get("search"):
            qs = qs.filter(
                Q(receipt_code__icontains=d["search"])
                | Q(church__name__icontains=d["search"])
                | Q(station__name__icontains=d["search"])
                | Q(agent__username__icontains=d["search"])
            )
        if d.get("company"):
            qs = qs.filter(station__company=d["company"])
        if d.get("station"):
            qs = qs.filter(station=d["station"])
        if d.get("status"):
            qs = qs.filter(status=d["status"])
        if d.get("date_from"):
            qs = qs.filter(created_at__date__gte=d["date_from"])
        if d.get("date_to"):
            qs = qs.filter(created_at__date__lte=d["date_to"])
    return qs


def _kpi_stats():
    today = timezone.now().date()
    this_month = timezone.now().replace(day=1, hour=0, minute=0, second=0)
    txs = Transaction.objects
    return {
        "today_levy": txs.filter(created_at__date=today).aggregate(t=Sum("levy_amount_usd"))["t"] or 0,
        "today_count": txs.filter(created_at__date=today).count(),
        "month_levy": txs.filter(created_at__gte=this_month).aggregate(t=Sum("levy_amount_usd"))["t"] or 0,
        "month_count": txs.filter(created_at__gte=this_month).count(),
        "total_levy": txs.aggregate(t=Sum("levy_amount_usd"))["t"] or 0,
        "total_count": txs.count(),
        "pending_count": txs.filter(status=Transaction.Status.PENDING).count(),
        "verified_count": txs.filter(status=Transaction.Status.VERIFIED).count(),
        "remitted_count": txs.filter(status=Transaction.Status.REMITTED).count(),
        "total_disbursed": Disbursement.objects.filter(status=Disbursement.Status.PAID).aggregate(t=Sum("amount_usd"))["t"] or 0,
        "pending_disburse": Disbursement.objects.filter(status=Disbursement.Status.SCHEDULED).count(),
        "by_company": list(
            txs.values("station__company__name", "station__company__id")
            .annotate(total=Sum("levy_amount_usd"), count=Count("id"))
            .order_by("-total")
        ),
        "by_fuel": list(
            txs.values("fuel_type__name")
            .annotate(total=Sum("levy_amount_usd"), count=Count("id"))
            .order_by("-total")
        ),
        "recent": txs.order_by("-created_at").select_related("station__company", "church", "agent")[:8],
    }


# ─── Dashboard ────────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    return render(request, "fuel/dashboard.html", {
        "stats": _kpi_stats(),
        "rate": get_usd_to_cdf_rate(),
        "companies": ParentCompany.objects.filter(is_active=True),
    })


@login_required
def dashboard_stats_partial(request):
    return render(request, "fuel/partials/stats_cards.html", {"stats": _kpi_stats()})


@login_required
def dashboard_chart_data(request):
    """JSON endpoint for Chart.js — last 30 days levy per day."""
    from datetime import timedelta
    today = timezone.now().date()
    data = []
    for i in range(29, -1, -1):
        d = today - timedelta(days=i)
        amt = Transaction.objects.filter(created_at__date=d).aggregate(t=Sum("levy_amount_usd"))["t"] or 0
        data.append({"date": d.strftime("%b %d"), "amount": float(amt)})
    return JsonResponse({"data": data})


# ─── Transactions ─────────────────────────────────────────────────────────────

@login_required
def transactions_list(request):
    qs = _tx_queryset(request)
    form = TransactionFilterForm(request.GET)
    totals = qs.aggregate(levy=Sum("levy_amount_usd"), count=Count("id"))
    ctx = {
        "transactions": qs[:300],
        "form": form,
        "totals": totals,
        "companies": ParentCompany.objects.filter(is_active=True),
        "stations": FuelStation.objects.filter(is_active=True).select_related("company"),
    }
    if _htmx(request):
        return render(request, "fuel/partials/tx_table.html", ctx)
    return render(request, "fuel/transactions/index.html", ctx)


@login_required
def transaction_detail(request, pk):
    tx = get_object_or_404(Transaction.objects.select_related(
        "station__company", "church", "agent", "fuel_type"
    ), pk=pk)
    audit_logs = tx.audit_logs.select_related("changed_by").all()
    form = TransactionStatusForm(instance=tx)
    return render(request, "fuel/transactions/detail.html", {
        "tx": tx, "audit_logs": audit_logs, "form": form,
    })


@login_required
@require_POST
def transaction_update_status(request, pk):
    tx = get_object_or_404(Transaction, pk=pk)
    form = TransactionStatusForm(request.POST, instance=tx)
    if form.is_valid():
        for field in ["status", "notes"]:
            if field in form.changed_data:
                record_audit_log(tx, request.user, field,
                                 getattr(tx, field), form.cleaned_data[field], request)
        form.save()
        messages.success(request, f"Transaction {tx.receipt_code} updated.")
        if _htmx(request):
            return render(request, "fuel/partials/tx_status_badge.html", {"tx": tx})
    return redirect("fuel:tx-detail", pk=pk)


@login_required
@require_POST
def transaction_bulk_action(request):
    ids = request.POST.getlist("selected")
    action = request.POST.get("action")
    STATUS_MAP = {"verify": Transaction.Status.VERIFIED, "remit": Transaction.Status.REMITTED}
    if action in STATUS_MAP and ids:
        qs = Transaction.objects.filter(pk__in=ids)
        for tx in qs:
            old = tx.status
            tx.status = STATUS_MAP[action]
            tx.save()
            record_audit_log(tx, request.user, "status", old, tx.status, request)
        messages.success(request, f"{qs.count()} transactions updated.")
    return redirect("fuel:transactions")


# ─── Companies ────────────────────────────────────────────────────────────────

@login_required
def company_list(request):
    qs = ParentCompany.objects.annotate(
        station_count=Count("stations", distinct=True),
        tx_count=Count("stations__transactions", distinct=True),
        total_levy=Sum("stations__transactions__levy_amount_usd"),
    ).order_by("name")
    return render(request, "fuel/companies/index.html", {"companies": qs})


@login_required
def company_create(request):
    form = ParentCompanyForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        company = form.save()
        messages.success(request, f'Company "{company.name}" created.')
        if _htmx(request):
            return render(request, "fuel/partials/company_row.html", {"company": company})
        return redirect("fuel:companies")
    return render(request, "fuel/companies/form.html", {"form": form, "title": "New Company"})


@login_required
def company_edit(request, pk):
    company = get_object_or_404(ParentCompany, pk=pk)
    form = ParentCompanyForm(request.POST or None, request.FILES or None, instance=company)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, f'Company "{company.name}" updated.')
        if _htmx(request):
            return render(request, "fuel/partials/company_row.html", {"company": company})
        return redirect("fuel:companies")
    return render(request, "fuel/companies/form.html", {"form": form, "company": company, "title": "Edit Company"})


@login_required
def company_detail(request, pk):
    company = get_object_or_404(ParentCompany, pk=pk)
    stations = company.stations.filter(is_active=True).annotate(
        tx_count=Count("transactions"), total_levy=Sum("transactions__levy_amount_usd")
    ).prefetch_related("churches")
    totals = Transaction.objects.filter(station__company=company).aggregate(
        total_usd=Sum("levy_amount_usd"), tx_count=Count("id")
    )
    if _htmx(request):
        return render(request, "fuel/partials/company_detail.html", {
            "company": company, "stations": stations, "totals": totals,
        })
    return render(request, "fuel/companies/detail.html", {
        "company": company, "stations": stations, "totals": totals,
    })


# ─── Stations ─────────────────────────────────────────────────────────────────

@login_required
def station_list(request):
    qs = FuelStation.objects.select_related("company").annotate(
        church_count=Count("churches", distinct=True),
        tx_count=Count("transactions", distinct=True),
        total_levy=Sum("transactions__levy_amount_usd"),
    ).order_by("company__name", "name")
    company_id = request.GET.get("company")
    if company_id:
        qs = qs.filter(company_id=company_id)
    return render(request, "fuel/stations/index.html", {
        "stations": qs,
        "companies": ParentCompany.objects.filter(is_active=True),
        "selected_company": company_id,
    })


@login_required
def station_create(request):
    form = FuelStationForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        station = form.save()
        messages.success(request, f'Station "{station.name}" created.')
        return redirect("fuel:stations")
    return render(request, "fuel/stations/form.html", {"form": form, "title": "New Station"})


@login_required
def station_edit(request, pk):
    station = get_object_or_404(FuelStation, pk=pk)
    form = FuelStationForm(request.POST or None, instance=station)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, f'Station "{station.name}" updated.')
        return redirect("fuel:stations")
    return render(request, "fuel/stations/form.html", {"form": form, "station": station, "title": "Edit Station"})


@login_required
def station_detail_view(request, pk):
    station = get_object_or_404(FuelStation.objects.select_related("company"), pk=pk)
    churches = station.churches.filter(is_active=True)
    recent_txs = station.transactions.select_related("church", "agent", "fuel_type")[:20]
    totals = station.transactions.aggregate(total_usd=Sum("levy_amount_usd"), tx_count=Count("id"))
    if _htmx(request):
        return render(request, "fuel/partials/station_detail.html", {
            "station": station, "churches": churches, "recent_txs": recent_txs, "totals": totals,
        })
    return render(request, "fuel/stations/detail.html", {
        "station": station, "churches": churches, "recent_txs": recent_txs, "totals": totals,
    })


# ─── Churches ─────────────────────────────────────────────────────────────────

@login_required
def church_list(request):
    qs = Church.objects.select_related("station__company").annotate(
        tx_count=Count("transactions"), total_levy=Sum("transactions__levy_amount_usd"),
        disburse_count=Count("disbursements"),
    ).order_by("station__company__name", "station__name", "name")
    return render(request, "fuel/churches/index.html", {"churches": qs})


@login_required
def church_create(request):
    form = ChurchForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        church = form.save()
        messages.success(request, f'Church "{church.name}" created.')
        return redirect("fuel:churches")
    return render(request, "fuel/churches/form.html", {"form": form, "title": "New Church"})


@login_required
def church_edit(request, pk):
    church = get_object_or_404(Church, pk=pk)
    form = ChurchForm(request.POST or None, instance=church)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, f'Church "{church.name}" updated.')
        return redirect("fuel:churches")
    return render(request, "fuel/churches/form.html", {"form": form, "church": church, "title": "Edit Church"})


@login_required
def church_detail(request, pk):
    church = get_object_or_404(Church.objects.select_related("station__company"), pk=pk)
    txs = church.transactions.select_related("agent", "fuel_type").order_by("-created_at")[:30]
    disbursements = church.disbursements.order_by("-created_at")[:10]
    totals = church.transactions.aggregate(levy=Sum("levy_amount_usd"), count=Count("id"))
    return render(request, "fuel/churches/detail.html", {
        "church": church, "txs": txs, "disbursements": disbursements, "totals": totals,
    })


# ─── Drivers (OSS registrations) ──────────────────────────────────────────────

def _driver_queryset(request):
    """Apply the Drivers-page search/filters from the query string."""
    qs = Driver.objects.all()
    cur = {key: request.GET.get(key, "").strip() for key in
           ("q", "commune", "vehicle_type", "fuel_type", "agent")}
    if cur["q"]:
        qs = qs.filter(
            Q(full_name__icontains=cur["q"]) | Q(phone__icontains=cur["q"])
            | Q(email__icontains=cur["q"]) | Q(quartier__icontains=cur["q"])
        )
    if cur["commune"]:
        qs = qs.filter(commune=cur["commune"])
    if cur["vehicle_type"]:
        qs = qs.filter(vehicle_type=cur["vehicle_type"])
    if cur["fuel_type"]:
        qs = qs.filter(fuel_type=cur["fuel_type"])
    if cur["agent"]:
        qs = qs.filter(field_agent=cur["agent"])
    return qs.order_by("full_name"), cur


@login_required
def driver_list(request):
    qs, cur = _driver_queryset(request)
    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get("page"))

    def _opts(field):
        return sorted(v for v in Driver.objects.values_list(field, flat=True).distinct() if v)

    params = request.GET.copy()
    params.pop("page", None)

    return render(request, "fuel/drivers/index.html", {
        "page_obj": page_obj,
        "total_count": Driver.objects.count(),
        "filtered_count": paginator.count,
        "communes": _opts("commune"),
        "vehicle_types": _opts("vehicle_type"),
        "fuel_types": _opts("fuel_type"),
        "agents": _opts("field_agent"),
        "cur": cur,
        "querystring": params.urlencode(),
    })


@login_required
def driver_detail(request, pk):
    driver = get_object_or_404(Driver, pk=pk)
    return render(request, "fuel/drivers/detail.html", {"driver": driver})


@login_required
def export_drivers_excel(request):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    qs, _ = _driver_queryset(request)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Drivers"
    hfill = PatternFill("solid", fgColor="1A3C5E")
    hfont = Font(bold=True, color="FFFFFF")
    headers = [
        "Full Name", "Phone", "Email", "Gender", "Marital Status", "Commune",
        "Quartier", "City/Country", "Vehicle Type", "Vehicle Color",
        "Litres/day", "Fuel Type", "Health Coverage", "Care Difficulty",
        "Dependents", "Field Agent", "Registered",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")

    def _yn(v):
        return "" if v is None else ("Yes" if v else "No")

    for row, d in enumerate(qs, 2):
        values = [
            d.full_name, d.phone, d.email, d.gender, d.marital_status, d.commune,
            d.quartier, d.city_country, d.vehicle_type, d.vehicle_color,
            d.daily_fuel_consumption, d.fuel_type, _yn(d.has_health_coverage),
            _yn(d.has_care_access_difficulty), d.dependents, d.field_agent,
            d.registration_date.strftime("%Y-%m-%d") if d.registration_date else "",
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row, column=col, value=val)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="LCI_drivers.xlsx"'
    wb.save(resp)
    return resp


# ─── Agents / Users ───────────────────────────────────────────────────────────

@login_required
def agent_list(request):
    users = User.objects.select_related("assigned_station__company", "managed_company").order_by("role", "username")
    return render(request, "fuel/agents/index.html", {"users": users})


@login_required
def agent_create(request):
    form = AgentForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = form.save()
        messages.success(request, f'User "{user.username}" created.')
        return redirect("fuel:agents")
    return render(request, "fuel/agents/form.html", {"form": form, "title": "New User"})


@login_required
def agent_edit(request, pk):
    user = get_object_or_404(User, pk=pk)
    form = AgentForm(request.POST or None, instance=user)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, f'User "{user.username}" updated.')
        return redirect("fuel:agents")
    return render(request, "fuel/agents/form.html", {"form": form, "user_obj": user, "title": "Edit User"})


# ─── Disbursements ────────────────────────────────────────────────────────────

@login_required
def disbursement_list(request):
    qs = Disbursement.objects.select_related("church__station__company", "prepared_by").order_by("-created_at")
    status_filter = request.GET.get("status")
    if status_filter:
        qs = qs.filter(status=status_filter)
    totals = qs.aggregate(total=Sum("amount_usd"), count=Count("id"))
    return render(request, "fuel/disbursements/index.html", {
        "disbursements": qs[:200],
        "totals": totals,
        "statuses": Disbursement.Status.choices,
        "status_filter": status_filter,
    })


@login_required
def disbursement_create(request):
    form = DisbursementForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        d = form.save(commit=False)
        d.prepared_by = request.user
        d.save()
        messages.success(request, f"Disbursement {d.reference} created.")
        return redirect("fuel:disbursements")
    return render(request, "fuel/disbursements/form.html", {"form": form, "title": "New Disbursement"})


@login_required
def disbursement_edit(request, pk):
    d = get_object_or_404(Disbursement, pk=pk)
    form = DisbursementForm(request.POST or None, instance=d)
    if request.method == "POST" and form.is_valid():
        form.save()
        if form.cleaned_data["status"] == Disbursement.Status.PAID and not d.paid_at:
            d.paid_at = timezone.now()
            d.save()
        messages.success(request, f"Disbursement {d.reference} updated.")
        return redirect("fuel:disbursements")
    return render(request, "fuel/disbursements/form.html", {"form": form, "disburse": d, "title": "Edit Disbursement"})


@login_required
@require_POST
def disbursement_mark_paid(request, pk):
    d = get_object_or_404(Disbursement, pk=pk)
    d.status = Disbursement.Status.PAID
    d.paid_at = timezone.now()
    d.save()
    messages.success(request, f"{d.reference} marked as paid.")
    if _htmx(request):
        return render(request, "fuel/partials/disburse_row.html", {"d": d})
    return redirect("fuel:disbursements")


# ─── Reports ──────────────────────────────────────────────────────────────────

@login_required
def reports(request):
    # Monthly levy roll-up for the past 12 months
    from datetime import date
    import calendar
    today = date.today()
    monthly = []
    for i in range(11, -1, -1):
        year = today.year
        month = today.month - i
        while month <= 0:
            month += 12
            year -= 1
        _, last_day = calendar.monthrange(year, month)
        period_start = date(year, month, 1)
        period_end = date(year, month, last_day)
        agg = Transaction.objects.filter(
            created_at__date__gte=period_start,
            created_at__date__lte=period_end,
        ).aggregate(levy=Sum("levy_amount_usd"), count=Count("id"))
        monthly.append({
            "label": period_start.strftime("%b %Y"),
            "levy": float(agg["levy"] or 0),
            "count": agg["count"] or 0,
        })

    # Per-church summary
    church_summary = list(
        Transaction.objects
        .values("church__name", "church__id", "church__station__name", "church__station__company__name")
        .annotate(total_levy=Sum("levy_amount_usd"), tx_count=Count("id"))
        .order_by("-total_levy")[:20]
    )

    # Per-fuel-type
    fuel_summary = list(
        Transaction.objects
        .values("fuel_type__name")
        .annotate(total=Sum("levy_amount_usd"), count=Count("id"))
        .order_by("-total")
    )

    return render(request, "fuel/reports/index.html", {
        "monthly": monthly,
        "church_summary": church_summary,
        "fuel_summary": fuel_summary,
        "stats": _kpi_stats(),
    })


# ─── Audit Log ────────────────────────────────────────────────────────────────

@login_required
def audit_log_list(request):
    qs = TransactionAuditLog.objects.select_related(
        "transaction__station__company", "changed_by"
    ).order_by("-changed_at")
    return render(request, "fuel/audit/index.html", {"logs": qs[:500]})


# ─── Fuel Types ───────────────────────────────────────────────────────────────

@login_required
def fuel_type_list(request):
    types = FuelType.objects.all()
    return render(request, "fuel/settings/fuel_types.html", {"types": types})


@login_required
def fuel_type_create(request):
    form = FuelTypeForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        ft = form.save()
        messages.success(request, f'Fuel type "{ft.name}" added.')
        if _htmx(request):
            return render(request, "fuel/partials/fuel_type_row.html", {"ft": ft})
        return redirect("fuel:fuel-types")
    if _htmx(request):
        return render(request, "fuel/partials/fuel_type_form.html", {"form": form})
    return render(request, "fuel/settings/fuel_types.html", {
        "types": FuelType.objects.all(), "form": form,
    })


@login_required
def fuel_type_edit(request, pk):
    ft = get_object_or_404(FuelType, pk=pk)
    form = FuelTypeForm(request.POST or None, instance=ft)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, f'Fuel type "{ft.name}" updated.')
        if _htmx(request):
            return render(request, "fuel/partials/fuel_type_row.html", {"ft": ft})
        return redirect("fuel:fuel-types")
    if _htmx(request):
        return render(request, "fuel/partials/fuel_type_form.html", {"form": form, "ft": ft})
    return render(request, "fuel/settings/fuel_types.html", {
        "types": FuelType.objects.all(), "form": form, "editing": ft,
    })


# ─── Verify (public) ──────────────────────────────────────────────────────────

def verify_receipt_page(request):
    code = request.GET.get("code", "").strip().upper()
    tx = None
    if code:
        tx = Transaction.objects.filter(receipt_code=code).select_related(
            "station__company", "church", "agent", "fuel_type"
        ).first()
    if _htmx(request):
        return render(request, "fuel/partials/verify_result.html", {"tx": tx, "code": code})
    return render(request, "fuel/verify.html", {"tx": tx, "code": code})


# ─── Export ───────────────────────────────────────────────────────────────────

@login_required
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    qs = _tx_queryset(request)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    hfill = PatternFill("solid", fgColor="1A3C5E")
    hfont = Font(bold=True, color="FFFFFF")
    headers = ["Receipt Code", "Date", "Company", "Station", "Church", "Agent",
               "Fuel Type", "Currency", "Amount USD", "Amount CDF", "Levy USD", "Levy CDF", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
    for row, tx in enumerate(qs, 2):
        ws.cell(row=row, column=1, value=tx.receipt_code)
        ws.cell(row=row, column=2, value=tx.created_at.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row, column=3, value=tx.station.company.name)
        ws.cell(row=row, column=4, value=tx.station.name)
        ws.cell(row=row, column=5, value=tx.church.name)
        ws.cell(row=row, column=6, value=tx.agent.username)
        ws.cell(row=row, column=7, value=tx.fuel_type.name)
        ws.cell(row=row, column=8, value=tx.currency_used)
        ws.cell(row=row, column=9, value=float(tx.amount_usd))
        ws.cell(row=row, column=10, value=float(tx.amount_cdf))
        ws.cell(row=row, column=11, value=float(tx.levy_amount_usd))
        ws.cell(row=row, column=12, value=float(tx.levy_amount_cdf))
        ws.cell(row=row, column=13, value=tx.get_status_display())
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="LCI_transactions.xlsx"'
    wb.save(resp)
    return resp


@login_required
def export_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    qs = _tx_queryset(request)
    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = 'attachment; filename="LCI_audit_report.pdf"'
    doc = SimpleDocTemplate(resp, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Lubumbashi Charity Fuel Initiative — Audit Report", styles["Title"]),
        Paragraph(f"Generated: {timezone.now():%Y-%m-%d %H:%M UTC}", styles["Normal"]),
        Spacer(1, 0.5*cm),
    ]
    table_data = [["Receipt", "Date", "Company", "Station", "Church", "Levy USD", "Status"]]
    for tx in qs[:500]:
        table_data.append([
            tx.receipt_code, tx.created_at.strftime("%Y-%m-%d"),
            tx.station.company.name[:20], tx.station.name[:20], tx.church.name[:20],
            f"${tx.levy_amount_usd:.2f}", tx.get_status_display(),
        ])
    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3C5E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF3F7")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    return resp


# ─── REST API ──────────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def api_currency_rate(request):
    return Response({"usd_to_cdf": str(get_usd_to_cdf_rate())})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def api_fuel_types(request):
    return Response(FuelTypeSerializer(FuelType.objects.filter(is_active=True), many=True).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def api_stations(request):
    return Response(FuelStationSerializer(
        FuelStation.objects.filter(is_active=True).select_related("company"), many=True
    ).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def api_churches(request):
    qs = Church.objects.filter(is_active=True).select_related("station__company")
    station_id = request.query_params.get("station")
    if station_id:
        qs = qs.filter(station_id=station_id)
    elif request.user.role == ROLE_STATION_AGENT and request.user.assigned_station:
        qs = qs.filter(station=request.user.assigned_station)
    return Response(ChurchSerializer(qs, many=True).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsStationAgent])
def api_transaction_create(request):
    serializer = TransactionCreateSerializer(data=request.data, context={"request": request})
    if serializer.is_valid():
        tx = serializer.save()
        return Response(TransactionSerializer(tx).data, status=drf_status.HTTP_201_CREATED)
    return Response(serializer.errors, status=drf_status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def api_transaction_list(request):
    qs = Transaction.objects.select_related("station__company", "church", "agent", "fuel_type")
    user = request.user
    if user.role == ROLE_STATION_AGENT:
        qs = qs.filter(station=user.assigned_station) if user.assigned_station else qs.none()
    elif user.role == ROLE_COMPANY_MANAGER:
        qs = qs.filter(station__company=user.managed_company) if user.managed_company else qs.none()
    return Response(TransactionSerializer(qs[:200], many=True).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def api_transaction_verify(request, receipt_code):
    tx = get_object_or_404(Transaction, receipt_code=receipt_code)
    return Response({
        "receipt_code": tx.receipt_code, "station": tx.station.name,
        "company": tx.station.company.name, "church": tx.church.name,
        "amount_usd": str(tx.amount_usd), "levy_usd": str(tx.levy_amount_usd),
        "status": tx.status, "created_at": tx.created_at, "valid": True,
    })


@api_view(["PATCH"])
@permission_classes([IsAuthenticated, IsNGOAdmin])
def api_transaction_status(request, pk):
    tx = get_object_or_404(Transaction, pk=pk)
    serializer = TransactionStatusUpdateSerializer(tx, data=request.data, partial=True)
    if serializer.is_valid():
        for field in ["status", "amount_usd", "amount_cdf"]:
            if field in serializer.validated_data:
                old_val = getattr(tx, field)
                new_val = serializer.validated_data[field]
                if str(old_val) != str(new_val):
                    record_audit_log(tx, request.user, field, old_val, new_val, request)
        serializer.save()
        return Response(TransactionSerializer(tx).data)
    return Response(serializer.errors, status=drf_status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsStationAgent])
def api_bulk_sync(request):
    results = []
    for tx_data in request.data.get("transactions", []):
        sync_id = tx_data.get("sync_id")
        if sync_id and Transaction.objects.filter(sync_id=sync_id).exists():
            results.append({"sync_id": sync_id, "status": "duplicate"})
            continue
        s = TransactionCreateSerializer(data=tx_data, context={"request": request})
        if s.is_valid():
            tx = s.save()
            results.append({"sync_id": sync_id, "receipt_code": tx.receipt_code, "levy_amount_usd": str(tx.levy_amount_usd), "status": "created"})
        else:
            results.append({"sync_id": sync_id, "errors": s.errors, "status": "error"})
    return Response({"results": results})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def api_audit_log(request, pk):
    tx = get_object_or_404(Transaction, pk=pk)
    return Response(TransactionAuditLogSerializer(tx.audit_logs.all(), many=True).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def api_profile(request):
    user = request.user
    return Response({
        "username": user.username,
        "email": user.email,
        "role": user.role,
        "assigned_station": str(user.assigned_station_id) if user.assigned_station_id else None,
        "managed_company": str(user.managed_company_id) if user.managed_company_id else None,
    })
