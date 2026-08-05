"""Currency and audit utilities for the fuel initiative."""
import re
from decimal import Decimal
from datetime import timedelta

import requests
from django.utils import timezone

# Row-level scoping now lives in fuel_app/scoping.py alongside the rules for
# every other model. Re-exported here because the mobile API imports it from
# this module.
from fuel_app.scoping import scope_transactions  # noqa: F401


RATE_CACHE_MINUTES = 30
FALLBACK_RATE = Decimal("2800.00")  # conservative fallback when API is unreachable


def normalize_phone(value) -> str:
    """Reduce a free-text phone to a comparable key: digits only, last 9.

    Driver phones come from a Google Form and vary wildly (spaces, +243
    country code, leading zeros). Keeping the trailing 9 digits makes a
    +243 / 0-prefixed / bare local number all compare equal, which is good
    enough for the loose driver↔transaction match.
    """
    if not value:
        return ""
    digits = re.sub(r"\D", "", str(value))
    return digits[-9:] if len(digits) >= 9 else digits


def get_usd_to_cdf_rate() -> Decimal:
    """
    Returns current USD→CDF rate.
    Checks DB cache first; fetches from open.er-api.com if stale.
    Falls back to last known rate or FALLBACK_RATE on failure.
    """
    from fuel_app.models import ExchangeRateCache

    cutoff = timezone.now() - timedelta(minutes=RATE_CACHE_MINUTES)
    cached = ExchangeRateCache.objects.filter(fetched_at__gte=cutoff).first()
    if cached:
        return cached.usd_to_cdf

    try:
        resp = requests.get(
            "https://open.er-api.com/v6/latest/USD",
            timeout=5,
        )
        resp.raise_for_status()
        data = resp.json()
        rate = Decimal(str(data["rates"]["CDF"]))
        ExchangeRateCache.objects.create(usd_to_cdf=rate, source="open.er-api.com")
        return rate
    except Exception:
        last = ExchangeRateCache.objects.order_by("-fetched_at").first()
        return last.usd_to_cdf if last else FALLBACK_RATE


def record_audit_log(transaction, user, field_name, old_value, new_value, request=None):
    """Write an immutable audit entry for a transaction field change."""
    from fuel_app.models import TransactionAuditLog

    ip = None
    if request:
        x_forwarded = request.META.get("HTTP_X_FORWARDED_FOR")
        ip = x_forwarded.split(",")[0].strip() if x_forwarded else request.META.get("REMOTE_ADDR")

    TransactionAuditLog.objects.create(
        transaction=transaction,
        changed_by=user,
        field_name=field_name,
        old_value=str(old_value) if old_value is not None else None,
        new_value=str(new_value) if new_value is not None else None,
        ip_address=ip,
    )


# ─── Dashboard / transaction / driver queryset helpers ─────────────────────
# Shared by the server-rendered dashboard (fuel_app/views.py) and the
# Next.js admin API (fuel_app/admin_views.py) so the business logic only
# lives in one place.

def kpi_stats(user=None):
    """Dashboard KPI block, scoped to what ``user`` may see.

    ``user=None`` means unscoped and is only for internal/management callers —
    every request-driven caller must pass ``request.user``.
    """
    from django.db.models import Count, Sum
    from fuel_app.models import Disbursement, Transaction
    from fuel_app.scoping import scope_disbursements

    today = timezone.now().date()
    this_month = timezone.now().replace(day=1, hour=0, minute=0, second=0)
    txs = Transaction.objects.all()
    disbursements = Disbursement.objects.all()
    if user is not None:
        txs = scope_transactions(user, txs)
        disbursements = scope_disbursements(user, disbursements)
    return {
        "today_levy": txs.filter(created_at__date=today).aggregate(t=Sum("levy_amount_usd"))["t"] or 0,
        "today_count": txs.filter(created_at__date=today).count(),
        "month_levy": txs.filter(created_at__gte=this_month).aggregate(t=Sum("levy_amount_usd"))["t"] or 0,
        "month_count": txs.filter(created_at__gte=this_month).count(),
        "total_levy": txs.aggregate(t=Sum("levy_amount_usd"))["t"] or 0,
        "total_count": txs.count(),
        "pending_count": txs.filter(status=Transaction.Status.PENDING).count(),
        "verified_count": txs.filter(status=Transaction.Status.VERIFIED).count(),
        "remitted_count": txs.filter(status=Transaction.Status.REMITTED).count(),
        "total_disbursed": disbursements.filter(status=Disbursement.Status.PAID).aggregate(t=Sum("amount_usd"))["t"] or 0,
        "pending_disburse": disbursements.filter(status=Disbursement.Status.SCHEDULED).count(),
        "by_company": list(
            txs.values("station__company__name", "station__company__id")
            .annotate(total=Sum("levy_amount_usd"), count=Count("id"))
            .order_by("-total")
        ),
        "by_fuel": list(
            txs.values("fuel_type__name")
            .annotate(total=Sum("levy_amount_usd"), count=Count("id"))
            .order_by("-total")
        ),
        "recent": txs.order_by("-created_at").select_related("station__company", "church", "agent")[:8],
    }


def top_stations_with_targets(user, limit=6):
    """Highest-earning stations this month, with progress against their target.

    Returns plain dicts so the same payload serves both the Django template and
    the JSON API. ``target_pct`` is ``None`` when no ``StationTarget`` exists
    for the current month, which the UI renders as "no target set" rather than
    an empty bar.
    """
    from django.db.models import F, Q, Sum
    from fuel_app.models import FuelStation, StationTarget
    from fuel_app.scoping import scope_stations

    today = timezone.now().date()
    month_start = today.replace(day=1)
    stations = list(
        scope_stations(user, FuelStation.objects.filter(is_active=True))
        .select_related("company")
        .annotate(month_levy=Sum(
            "transactions__levy_amount_usd",
            filter=Q(transactions__created_at__date__gte=month_start),
        ))
        .order_by(F("month_levy").desc(nulls_last=True))[:limit]
    )
    targets = {
        t.station_id: t.target_usd
        for t in StationTarget.objects.filter(year=today.year, month=today.month)
    }
    rows = []
    for s in stations:
        month_levy = s.month_levy or 0
        target = targets.get(s.id)
        rows.append({
            "id": str(s.id),
            "name": s.name,
            "company_name": s.company.name,
            "month_levy": float(month_levy),
            "target_usd": float(target) if target else None,
            "target_pct": (
                min(100, round(float(month_levy) / float(target) * 100))
                if target else None
            ),
        })
    return rows


def levy_by_day(user, days=30):
    """Daily levy totals for the trend chart, as one grouped query.

    Days with no transactions still appear with ``0`` so the line has no gaps.
    """
    from datetime import timedelta

    from django.db.models import Sum
    from django.db.models.functions import TruncDate
    from fuel_app.models import Transaction
    from fuel_app.scoping import scope_transactions

    days = max(2, min(int(days), 365))
    today = timezone.now().date()
    start = today - timedelta(days=days - 1)

    totals = {
        row["day"]: float(row["total"] or 0)
        for row in scope_transactions(user, Transaction.objects.all())
        .filter(created_at__date__gte=start)
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(total=Sum("levy_amount_usd"))
    }
    return [
        {
            "date": (d := start + timedelta(days=i)).strftime("%d %b"),
            "amount": totals.get(d, 0.0),
        }
        for i in range(days)
    ]


def tx_queryset(request):
    """Filtered transaction list for the current request, scoped to the caller.

    Scoping is applied *before* the filter form so that a manager or agent
    cannot widen their view by passing an out-of-scope ``?company=`` or
    ``?station=`` — those filters can only ever narrow what is already visible.
    """
    from django.db.models import Q
    from fuel_app.forms import TransactionFilterForm
    from fuel_app.models import Transaction

    qs = scope_transactions(request.user, Transaction.objects.select_related(
        "station__company", "church", "agent", "fuel_type"
    ))
    f = TransactionFilterForm(request.GET)
    if f.is_valid():
        d = f.cleaned_data
        if d.get("search"):
            qs = qs.filter(
                Q(receipt_code__icontains=d["search"])
                | Q(church__name__icontains=d["search"])
                | Q(station__name__icontains=d["search"])
                | Q(agent__username__icontains=d["search"])
            )
        if d.get("company"):
            qs = qs.filter(station__company=d["company"])
        if d.get("station"):
            qs = qs.filter(station=d["station"])
        if d.get("status"):
            qs = qs.filter(status=d["status"])
        if d.get("date_from"):
            qs = qs.filter(created_at__date__gte=d["date_from"])
        if d.get("date_to"):
            qs = qs.filter(created_at__date__lte=d["date_to"])
    return qs


# ─── Transaction history (mobile /api/…/history/ endpoints) ────────────────

def filter_history(qs, params):
    """Apply the shared ``?from=&to=&status=&church=&fuel_type=`` filters.

    Dates are ISO ``YYYY-MM-DD`` and inclusive on both ends; an unparseable
    value is ignored rather than erroring, so a stale mobile client sending a
    bad filter still gets its history back.
    """
    from datetime import date

    def _parse(key):
        raw = (params.get(key) or "").strip()
        try:
            return date.fromisoformat(raw) if raw else None
        except ValueError:
            return None

    date_from, date_to = _parse("from"), _parse("to")
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    status = (params.get("status") or "").strip().upper()
    if status:
        qs = qs.filter(status=status)
    for key, field in (("church", "church_id"), ("fuel_type", "fuel_type_id")):
        value = (params.get(key) or "").strip()
        if value:
            qs = qs.filter(**{field: value})
    return qs


def money_str(value, places="0.01"):
    """Format an aggregate as a fixed-precision string.

    SQLite's SUM() comes back float-backed, so a plain str() yields noise like
    "0.430000000000000". Quantize to the column's own precision instead.
    """
    return str(Decimal(str(value or 0)).quantize(Decimal(places)))


def history_summary(qs):
    """Totals + per-status breakdown for a history queryset (2 queries)."""
    from django.db.models import Count, Max, Min, Sum

    agg = qs.aggregate(
        count=Count("id"),
        total_amount_usd=Sum("amount_usd"),
        total_amount_cdf=Sum("amount_cdf"),
        total_levy_usd=Sum("levy_amount_usd"),
        total_levy_cdf=Sum("levy_amount_cdf"),
        first_at=Min("created_at"),
        last_at=Max("created_at"),
    )
    by_status = {
        row["status"]: {"count": row["count"], "levy_usd": money_str(row["levy"], "0.0001")}
        for row in qs.values("status").annotate(count=Count("id"), levy=Sum("levy_amount_usd"))
    }
    return {
        "count": agg["count"] or 0,
        "total_amount_usd": money_str(agg["total_amount_usd"]),
        "total_amount_cdf": money_str(agg["total_amount_cdf"]),
        "total_levy_usd": money_str(agg["total_levy_usd"], "0.0001"),
        "total_levy_cdf": money_str(agg["total_levy_cdf"], "0.0001"),
        "first_at": agg["first_at"],
        "last_at": agg["last_at"],
        "by_status": by_status,
    }


def paginate_history(qs, request, serializer_class):
    """Page a history queryset with the project-wide PageNumberPagination.

    ``@api_view`` functions don't get DRF's automatic pagination, so the
    paginator is driven by hand here and reused by both history endpoints.
    """
    from rest_framework.pagination import PageNumberPagination

    paginator = PageNumberPagination()
    paginator.page_size_query_param = "page_size"
    paginator.max_page_size = 200
    page = paginator.paginate_queryset(qs, request)
    return {
        "count": paginator.page.paginator.count,
        "page": paginator.page.number,
        "num_pages": paginator.page.paginator.num_pages,
        "next": paginator.get_next_link(),
        "previous": paginator.get_previous_link(),
        "results": serializer_class(page, many=True).data,
    }


DRIVER_SORTABLE = {
    "name": "full_name", "commune": "commune", "vehicle": "vehicle_type",
    "fuel": "fuel_type", "consumption": "daily_fuel_consumption",
    "agent": "field_agent", "registered": "registration_date",
}
CONSUMPTION_ORDER = ["1 à 4", "5 à 10", "11 à 20", "21 à 30", "31 à 45", "Autre"]


def driver_queryset(request):
    """Apply the Drivers-page search/filters from the query string."""
    from django.db.models import Q
    from fuel_app.models import Driver

    qs = Driver.objects.all()
    cur = {key: request.GET.get(key, "").strip() for key in
           ("q", "commune", "vehicle_type", "fuel_type", "agent")}
    if cur["q"]:
        qs = qs.filter(
            Q(full_name__icontains=cur["q"]) | Q(phone__icontains=cur["q"])
            | Q(email__icontains=cur["q"]) | Q(quartier__icontains=cur["q"])
        )
    if cur["commune"]:
        qs = qs.filter(commune=cur["commune"])
    if cur["vehicle_type"]:
        qs = qs.filter(vehicle_type=cur["vehicle_type"])
    if cur["fuel_type"]:
        qs = qs.filter(fuel_type=cur["fuel_type"])
    if cur["agent"]:
        qs = qs.filter(field_agent=cur["agent"])
    return qs.order_by("full_name"), cur


def driver_transactions(driver, user):
    """A driver's levy history, scoped to what ``user`` may see.

    There is no FK from Transaction to Driver — transactions carry a
    ``driver_phone`` string written by the mobile app. Both sides are reduced
    with ``normalize_phone`` so a ``+243``-prefixed, ``0``-prefixed or bare
    local number all match. A driver with no phone on file matches nothing.
    """
    from fuel_app.models import Transaction
    from fuel_app.scoping import scope_transactions

    phone = normalize_phone(driver.phone)
    if not phone:
        return Transaction.objects.none()
    return scope_transactions(user, Transaction.objects.filter(driver_phone=phone)
                              .select_related("station__company", "church", "fuel_type", "agent")
                              .order_by("-created_at"))


def qr_data_uri(data):
    """Render `data` as a QR code and return it as a base64 PNG data URI."""
    import base64
    import io

    import qrcode

    img = qrcode.make(data)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    encoded = base64.b64encode(buf.getvalue()).decode("ascii")
    return f"data:image/png;base64,{encoded}"


def barcode_data_uri(code):
    """Render `code` as a Code128 barcode and return it as a base64 SVG data URI."""
    import base64
    import io

    import barcode
    from barcode.writer import SVGWriter

    cls = barcode.get_barcode_class("code128")
    buf = io.BytesIO()
    cls(code, writer=SVGWriter()).write(
        buf,
        options={"write_text": False, "module_height": 12.0, "quiet_zone": 1.0},
    )
    encoded = base64.b64encode(buf.getvalue()).decode("ascii")
    return f"data:image/svg+xml;base64,{encoded}"


def driver_card_number(driver):
    """Stable 16-digit card number derived from the driver's UUID.

    Returns ``(raw_digits, grouped)`` where ``grouped`` matches the card layout
    (4-3-3-3-3, e.g. ``0001 004 550 650 111``).
    """
    raw = f"{driver.pk.int % 10 ** 16:016d}"
    grouped = f"{raw[0:4]} {raw[4:7]} {raw[7:10]} {raw[10:13]} {raw[13:16]}"
    return raw, grouped


def image_data_uri(path):
    """Read an image file and return it as a base64 data URI (empty string if missing)."""
    import base64
    import mimetypes

    try:
        data = path.read_bytes()
    except OSError:
        return ""
    mime = mimetypes.guess_type(str(path))[0] or "image/jpeg"
    encoded = base64.b64encode(data).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def build_transactions_excel(qs):
    """Build an .xlsx HttpResponse of the given Transaction queryset."""
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    hfill = PatternFill("solid", fgColor="1A3C5E")
    hfont = Font(bold=True, color="FFFFFF")
    headers = ["Receipt Code", "Date", "Company", "Station", "Church", "Agent",
               "Fuel Type", "Currency", "Amount USD", "Amount CDF", "Levy USD", "Levy CDF", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
    for row, tx in enumerate(qs, 2):
        ws.cell(row=row, column=1, value=tx.receipt_code)
        ws.cell(row=row, column=2, value=tx.created_at.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row, column=3, value=tx.station.company.name)
        ws.cell(row=row, column=4, value=tx.station.name)
        ws.cell(row=row, column=5, value=tx.church.name)
        ws.cell(row=row, column=6, value=tx.agent.username)
        ws.cell(row=row, column=7, value=tx.fuel_type.name)
        ws.cell(row=row, column=8, value=tx.currency_used)
        ws.cell(row=row, column=9, value=float(tx.amount_usd))
        ws.cell(row=row, column=10, value=float(tx.amount_cdf))
        ws.cell(row=row, column=11, value=float(tx.levy_amount_usd))
        ws.cell(row=row, column=12, value=float(tx.levy_amount_cdf))
        ws.cell(row=row, column=13, value=tx.get_status_display())
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="LCI_transactions.xlsx"'
    wb.save(resp)
    return resp


def build_transactions_pdf(qs):
    """Build a PDF HttpResponse audit report of the given Transaction queryset."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from django.http import HttpResponse

    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = 'attachment; filename="LCI_audit_report.pdf"'
    doc = SimpleDocTemplate(resp, pagesize=landscape(A4), leftMargin=1 * cm, rightMargin=1 * cm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Lubumbashi Charity Fuel Initiative — Audit Report", styles["Title"]),
        Paragraph(f"Generated: {timezone.now():%Y-%m-%d %H:%M UTC}", styles["Normal"]),
        Spacer(1, 0.5 * cm),
    ]
    table_data = [["Receipt", "Date", "Company", "Station", "Church", "Levy USD", "Status"]]
    for tx in qs[:500]:
        table_data.append([
            tx.receipt_code, tx.created_at.strftime("%Y-%m-%d"),
            tx.station.company.name[:20], tx.station.name[:20], tx.church.name[:20],
            f"${tx.levy_amount_usd:.2f}", tx.get_status_display(),
        ])
    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3C5E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF3F7")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    return resp


def build_drivers_excel(qs):
    """Build an .xlsx HttpResponse of the given Driver queryset."""
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Drivers"
    hfill = PatternFill("solid", fgColor="1A3C5E")
    hfont = Font(bold=True, color="FFFFFF")
    headers = [
        "Full Name", "Phone", "Email", "Gender", "Marital Status", "Commune",
        "Quartier", "City/Country", "Vehicle Type", "Vehicle Color",
        "Litres/day", "Fuel Type", "Health Coverage", "Care Difficulty",
        "Dependents", "Field Agent", "Registered",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")

    def _yn(v):
        return "" if v is None else ("Yes" if v else "No")

    for row, d in enumerate(qs, 2):
        values = [
            d.full_name, d.phone, d.email, d.gender, d.marital_status, d.commune,
            d.quartier, d.city_country, d.vehicle_type, d.vehicle_color,
            d.daily_fuel_consumption, d.fuel_type, _yn(d.has_health_coverage),
            _yn(d.has_care_access_difficulty), d.dependents, d.field_agent,
            d.registration_date.strftime("%Y-%m-%d") if d.registration_date else "",
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row, column=col, value=val)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="LCI_drivers.xlsx"'
    wb.save(resp)
    return resp
